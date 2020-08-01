#  pip install openpyxl
# pipenv install openpyxl

import openpyxl

# wb = openpyxl.Workbook() # Can create empty workbook object in the memory or load an existing workbook - HERE WE HAVVE EMPTY WORKBOOK

# For loading an exisiting workbook
wb = openpyxl.load_workbook("analysis.xlsx")

print(wb.sheetnames)  # Gives names of the sheets present in the excel file

sheet = wb["Analysis Q4"]  # Gives sheet object

# Will create a new sheet named as Sheet 1 with index 0 i.e will be placed before exisitng sheet or the first sheet
wb.create_sheet("Sheet 1", 0)
# wb.remove_sheet(sheet) # Will remove the sheet object

# ACCESSING INDIVIDUAL CELL ORRANGE OF CELLS

cell = sheet["a1"]  # Passing the coordinates of a cell
print("\n", cell.value)

# cell.value = "product_id"  # Can also change thevalue of a cell
print(cell.row)
print(cell.column)
print(cell.coordinate, "\n")

# ANOTHER METHOD TO ACCESS CELL

# sheet.cell(row=1, column=1)  # Passing Arguments (row, column)

print(sheet.max_row)  # Number of rows
print(sheet.max_column)  # Number of columns

# ITERATING OUR CELLS

# sheet.max_row returns a number and range function gives a number but starts from o and hence indexed with 1, sheet.max_row+1 to accesss the last row
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)  # Value of each cell
    print("\n")

# sheet["a"]  # Will return all the values in the "A" column
# sheet["a:c"]  # Will return range of column values
# sheet["a1:c3"]  # Setting Coordintes will return all the cell values from a1 to c3
# sheet[1:3] # Will return all the cell value in row 1 to 3

# Is to add a row at the end and passing list of values
sheet.append([1, "KFFL", 320])
# sheet.insert_rows()  # Will insert row at the given index number
# sheet.insert_cols()  # Inserting a column at a index number
# sheet.delete_rows()  # Deleting rows at index , similar for columns

# SAVING WORKBOOK IN THE FILE
wb.save("analysis2.xlsx")
